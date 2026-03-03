from fastapi import FastAPI, Request
from fastapi.responses import StreamingResponse
from fastapi.templating import Jinja2Templates
from pydantic import BaseModel
from tcgdexsdk import Query, TCGdex
import aiohttp
import openpyxl
from openpyxl.drawing.image import Image
from PIL import Image as PILImage
import io
import asyncio
import os

app = FastAPI()
tcgdex = TCGdex()
templates = Jinja2Templates(directory="templates")


class FiltroRequest(BaseModel):
    tipo: str
    valor: str


# 🔹 ROTA PRINCIPAL (AGORA NÃO DÁ MAIS 404)
@app.get("/")
async def home(request: Request):
    return templates.TemplateResponse("index.html", {"request": request})


async def buscar_cartas(tipo, valor):
    query = Query().paginate(page=1, itemsPerPage=1000)

    if tipo == "illustrator":
        query = query.contains("illustrator", valor)
    elif tipo == "set":
        query = query.contains("set.name", valor)
    elif tipo == "series":
        query = query.contains("series.name", valor)
    elif tipo == "name":
        query = query.contains("name", valor)

    return await tcgdex.card.list(query)

async def info_cartas(card_id):
    return await tcgdex.card.get(card_id)

async def download_image(session, url):
    async with session.get(url) as response:
        if response.status == 200:
            return await response.read()
        return None


def cortar_ilustracao(pil_img):
    largura, altura = pil_img.size

    esquerda = int(largura * 0.08)
    direita = int(largura * 0.92)
    topo = int(altura * 0.12)
    base = int(altura * 0.48)

    return pil_img.crop((esquerda, topo, direita, base))


@app.post("/buscar")
async def gerar_excel(filtro: FiltroRequest):

    pokemons = await buscar_cartas(filtro.tipo, filtro.valor)

    wb = openpyxl.Workbook()
    ws = wb.active
    
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 30

    async with aiohttp.ClientSession() as session:

        semaphore = asyncio.Semaphore(20)

        async def get_full_info(pokemon):
            async with semaphore:
                return await tcgdex.card.get(pokemon.id)

        async def get_image(pokemon):
            if pokemon.image:
                async with semaphore:
                    return await download_image(session, f"{pokemon.image}/low.png")
            return None

        info_tasks = [get_full_info(p) for p in pokemons if p.id]
        image_tasks = [get_image(p) for p in pokemons]

        infos = await asyncio.gather(*info_tasks)
        images = await asyncio.gather(*image_tasks)

        row = 1
        info_index = 0

        for i, pokemon in enumerate(pokemons):
            
            ws.row_dimensions[row].height = 100
            ws.cell(row=row, column=2).value = pokemon.name

            if pokemon.id:
                ws.cell(row=row, column=4).value = infos[info_index].set.name
                ws.cell(row=row, column=3).value = f'{pokemon.localId}/{infos[info_index].set.cardCount.total}'
                info_index += 1


            img_data = images[i]
            if img_data:
                pil_img = PILImage.open(io.BytesIO(img_data))
                pil_img = cortar_ilustracao(pil_img)
                pil_img.thumbnail((300, 300))

                temp_bytes = io.BytesIO()
                pil_img.save(temp_bytes, format="PNG")
                temp_bytes.seek(0)

                img = Image(temp_bytes)
                ws.add_image(img, f'A{row}')

            row += 1

    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename=checklist.xlsx"
        }
    )
