from tcgdexsdk import Query, TCGdex
import asyncio
import aiohttp
import openpyxl
from openpyxl.drawing.image import Image
from PIL import Image as PILImage
import io

tcgdex = TCGdex()

async def lista():
    return await tcgdex.card.list(
        Query()
        .contains("illustrator", "Komiya")
        .paginate(page=1, itemsPerPage=1000)
    )

async def download_image(session, url):
    try:
        async with session.get(url) as response:
            if response.status == 200:
                return await response.read()
            else:
                print(f"Erro {response.status} ao baixar {url}")
                return None
    except Exception as e:
        print(f"Erro ao baixar {url}: {e}")
        return None

def cortar_ilustracao(pil_img):
    largura, altura = pil_img.size

    esquerda = int(largura * 0.08)
    direita = int(largura * 0.92)
    topo = int(altura * 0.12)
    base = int(altura * 0.48)

    return pil_img.crop((esquerda, topo, direita, base))

async def main():
    pokemons = await lista()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 30

    async with aiohttp.ClientSession() as session:

        row = 1

        for pokemon in pokemons:
            try:
                ws.cell(row=row, column=2).value = pokemon.name

                if not pokemon.image:
                    ws.cell(row=row, column=1).value = "Sem imagem disponível"
                    print(f"{pokemon.name} não possui imagem")
                else:
                    url = f"{pokemon.image}/low.png"
                    img_data = await download_image(session, url)

                    if img_data is None:
                        ws.cell(row=row, column=1).value = "Erro ao baixar imagem"
                    else:
                        image_file = io.BytesIO(img_data)
                        pil_img = PILImage.open(image_file)

                        pil_img = cortar_ilustracao(pil_img)
                        pil_img.thumbnail((300, 300))

                        temp_bytes = io.BytesIO()
                        pil_img.save(temp_bytes, format="PNG")
                        temp_bytes.seek(0)

                        img = Image(temp_bytes)
                        ws.row_dimensions[row].height = 100
                        ws.add_image(img, f'A{row}')

                print(f"Adicionado: {pokemon.name}")
                row += 1

            except Exception as e:
                ws.cell(row=row, column=1).value = "Erro ao processar"
                ws.cell(row=row, column=2).value = pokemon.name
                print(f"Erro ao processar {pokemon.name}: {e}")
                row += 1

    wb.save("C:\\Users\\giova\\Desktop\\pokemons.xlsx")
    print("Arquivo salvo com sucesso!")

asyncio.run(main())